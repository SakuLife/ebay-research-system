"""
Auto Research結果のレポートExcel生成ツール.

スプレッドシートの入力シートから指定行以降のデータを読み取り、
分析・確認用のExcelレポートを生成する。
"""

import os
import sys
from datetime import datetime
from pathlib import Path

# プロジェクトルートをパスに追加
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / ".env")

import gspread
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_sheets_client():
    """Google Sheetsクライアントを取得."""
    cred_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if os.path.isfile(cred_path):
        gc = gspread.service_account(filename=cred_path)
    else:
        import json
        cred_data = json.loads(cred_path)
        gc = gspread.service_account_from_dict(cred_data)

    spreadsheet_id = os.getenv("SHEETS_SPREADSHEET_ID", "")
    if "spreadsheets/d/" in spreadsheet_id:
        spreadsheet_id = spreadsheet_id.split("spreadsheets/d/")[1].split("/")[0]
    return gc.open_by_key(spreadsheet_id)


def fetch_data(start_row: int) -> list[list[str]]:
    """入力シートから指定行以降のデータを取得."""
    spreadsheet = get_sheets_client()
    ws = spreadsheet.worksheet("入力シート")
    all_data = ws.get_all_values()

    # start_rowは1-indexed（ヘッダー=1行目）
    rows = all_data[start_row - 1:]  # 0-indexed
    # 空行を除外
    return [r for r in rows if any(cell.strip() for cell in r)]


def create_report(data: list[list[str]], output_path: str, start_row: int):
    """Excelレポートを生成."""
    wb = Workbook()

    # === Sheet 1: 商品一覧 ===
    ws1 = wb.active
    ws1.title = "商品一覧"

    # ヘッダー
    headers = [
        "No.",
        "行番号",
        "日付",
        "キーワード",
        "カテゴリ",
        "新品/中古",
        "eBayリンク",
        "販売数",
        "販売価格($)",
        "販売送料($)",
        "仕入先① 商品名",
        "仕入先① リンク",
        "仕入先① 価格(円)",
        "仕入先② 商品名",
        "仕入先② リンク",
        "仕入先② 価格(円)",
        "還付抜き利益(円)",
        "利益率%(還付抜き)",
        "還付あり利益(円)",
        "利益率%(還付あり)",
        "ステータス",
        "メモ",
        # セラーページ確認用（手入力）
        "【手入力】セラーページ商品数",
        "【手入力】セラーページ最安値($)",
        "【手入力】確認結果",
        "【手入力】備考",
    ]

    # ヘッダースタイル
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    manual_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    manual_font = Font(bold=True, size=10, color="8B4513")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if header.startswith("【手入力】"):
            cell.fill = manual_fill
            cell.font = manual_font
        else:
            cell.fill = header_fill
            cell.font = header_font

    # データ行
    # スプレッドシートの列マッピング（0-indexed）
    # A=0:日付, B=1:キーワード, C=2:カテゴリ, D=3:カテゴリ番号, E=4:新品中古
    # F=5:国内最安①商品名, G=6:①リンク, H=7:①価格
    # I=8:②商品名, J=9:②リンク, K=10:②価格
    # L=11:③商品名, M=12:③リンク, N=13:③価格
    # O=14:eBayリンク, P=15:販売数, Q=16:販売価格, R=17:販売送料
    # S=18:還付抜き利益, T=19:利益率%(還付抜き), U=20:還付あり利益, V=21:利益率%(還付あり)
    # W=22:ステータス, X=23:出品フラグ, Y=24:メモ

    profit_ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    profit_ng_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    link_font = Font(color="0563C1", underline="single", size=9)
    normal_font = Font(size=9)

    for row_idx, row_data in enumerate(data):
        excel_row = row_idx + 2  # Excelの行（ヘッダー+1）
        sheet_row = start_row + row_idx  # スプレッドシートの行番号

        def safe_get(idx: int) -> str:
            return row_data[idx].strip() if idx < len(row_data) else ""

        # 利益判定
        profit_str = safe_get(18)
        try:
            profit_val = float(profit_str.replace(",", ""))
            is_profitable = profit_val >= 5000
        except (ValueError, AttributeError):
            profit_val = None
            is_profitable = False

        row_fill = profit_ok_fill if is_profitable else None

        values = [
            row_idx + 1,          # No.
            sheet_row,            # 行番号
            safe_get(0),          # 日付
            safe_get(1),          # キーワード
            safe_get(2),          # カテゴリ
            safe_get(4),          # 新品/中古
            safe_get(14),         # eBayリンク
            safe_get(15),         # 販売数
            safe_get(16),         # 販売価格
            safe_get(17),         # 販売送料
            safe_get(5),          # 仕入先① 商品名
            safe_get(6),          # 仕入先① リンク
            safe_get(7),          # 仕入先① 価格
            safe_get(8),          # 仕入先② 商品名
            safe_get(9),          # 仕入先② リンク
            safe_get(10),         # 仕入先② 価格
            safe_get(18),         # 還付抜き利益
            safe_get(19),         # 利益率%(還付抜き)
            safe_get(20),         # 還付あり利益
            safe_get(21),         # 利益率%(還付あり)
            safe_get(22),         # ステータス
            safe_get(24),         # メモ
            "",                   # セラーページ商品数（手入力）
            "",                   # セラーページ最安値（手入力）
            "",                   # 確認結果（手入力）
            "",                   # 備考（手入力）
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=excel_row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border

            # リンク列はハイパーリンク化
            if col_idx in (7, 12, 15) and val and val.startswith("http"):
                cell.font = link_font
                cell.hyperlink = val
                cell.value = val[:60] + "..." if len(val) > 60 else val

            # 利益OKの行は緑背景
            if row_fill and col_idx <= 22:
                cell.fill = row_fill
            # 利益NGの行で利益列は赤背景
            if not is_profitable and profit_val is not None and col_idx in (17, 18):
                cell.fill = profit_ng_fill

            # 手入力列は黄色背景
            if col_idx >= 23:
                cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

    # 列幅調整
    col_widths = {
        1: 5, 2: 6, 3: 11, 4: 15, 5: 20, 6: 8,
        7: 25, 8: 8, 9: 10, 10: 10,
        11: 25, 12: 25, 13: 10,
        14: 25, 15: 25, 16: 10,
        17: 12, 18: 10, 19: 12, 20: 10,
        21: 10, 22: 25,
        23: 15, 24: 15, 25: 12, 26: 20,
    }
    for col, width in col_widths.items():
        ws1.column_dimensions[get_column_letter(col)].width = width

    # フィルタ設定
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    # ウィンドウ枠の固定（ヘッダー行 + 左3列）
    ws1.freeze_panes = "D2"

    # === Sheet 2: サマリー ===
    ws2 = wb.create_sheet("サマリー")

    summary_header_font = Font(bold=True, size=11)
    summary_items = [
        ("レポート生成日時", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("対象行", f"{start_row}行目〜{start_row + len(data) - 1}行目（{len(data)}件）"),
        ("", ""),
        ("--- キーワード別集計 ---", ""),
    ]

    # キーワード別集計
    keyword_stats: dict[str, dict] = {}
    for row_data in data:
        kw = row_data[1].strip() if len(row_data) > 1 else "(不明)"
        if not kw:
            kw = "(不明)"
        if kw not in keyword_stats:
            keyword_stats[kw] = {"count": 0, "profitable": 0, "total_profit": 0}
        keyword_stats[kw]["count"] += 1
        try:
            profit = float(row_data[18].replace(",", "")) if len(row_data) > 18 and row_data[18].strip() else 0
            if profit >= 5000:
                keyword_stats[kw]["profitable"] += 1
            keyword_stats[kw]["total_profit"] += profit
        except (ValueError, IndexError):
            pass

    for kw, stats in keyword_stats.items():
        summary_items.append((
            kw,
            f"{stats['count']}件 (利益OK: {stats['profitable']}件, 合計利益: JPY {stats['total_profit']:,.0f})"
        ))

    summary_items.extend([
        ("", ""),
        ("--- 全体 ---", ""),
        ("総件数", f"{len(data)}件"),
        ("利益5000円以上", f"{sum(s['profitable'] for s in keyword_stats.values())}件"),
        ("合計利益(還付抜き)", f"JPY {sum(s['total_profit'] for s in keyword_stats.values()):,.0f}"),
    ])

    for row_idx, (label, value) in enumerate(summary_items, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if label.startswith("---"):
            cell_a.font = summary_header_font
        cell_a.font = Font(bold=True, size=10)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50

    # 保存
    wb.save(output_path)
    print(f"レポート生成完了: {output_path}")
    print(f"  対象: {len(data)}件 (行{start_row}〜{start_row + len(data) - 1})")
    print(f"  利益5000円以上: {sum(s['profitable'] for s in keyword_stats.values())}件")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Auto Research結果レポート生成")
    parser.add_argument("--start-row", type=int, default=664,
                        help="スプレッドシートの開始行番号 (default: 664)")
    parser.add_argument("--output", type=str, default="",
                        help="出力ファイルパス (default: reports/report_YYYYMMDD.xlsx)")
    args = parser.parse_args()

    # 出力パス
    if args.output:
        output_path = args.output
    else:
        reports_dir = Path(__file__).resolve().parent.parent / "reports"
        reports_dir.mkdir(exist_ok=True)
        output_path = str(reports_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

    print(f"スプレッドシートからデータ取得中（{args.start_row}行目〜）...")
    data = fetch_data(args.start_row)

    if not data:
        print("データが見つかりません")
        return

    create_report(data, output_path, args.start_row)


if __name__ == "__main__":
    main()
